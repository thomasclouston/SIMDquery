from openpyxl import Workbook, load_workbook
import pandas as pd
#CSV input file:
df = pd.read_csv("/Users/Thomas/Documents/py/SIMDQuerytool/SIMDQuery2016/QueryPostcodes.csv")
#number of postcodes entered output
number_of_postcodes = int(len(df))
print ("Number of postcodes queried: ", number_of_postcodes)
print("Do NOT press any key...")
#change input csv to list
postcode_list= df.values.tolist()
#SIMD_2020_Postcode_lookup_tool return
path = "/Users/Thomas/Documents/py/SIMDQuerytool/SIMDQuery2016/2016Postcodes.xlsx"
wb = load_workbook(path)
ws= wb['All postcodes']
postcodes=[]
datazone_lists=[]
datazone_lists.append("datazone_lists")
postcodes.append("Postcode")
 # logic
for i in range(1,ws.max_row):
    for postcode in postcode_list:
       if ws.cell(row=i, column=1).value == postcode[0]:
          for j in range(1,2):
              postcodes.append (str(ws.cell(row=i, column=j).value))
          for j in range(2,3):
              datazone_lists.append (str(ws.cell(row=i, column=j).value))

#Background generating to ask user what they wish to investigate
enum_list_dict = []
list_of_factors=['council_area_list','total_population','Working_Age_population','Income_rate',\
'Income_count','Employment_rate','Employment_count','CIF','ALCOHOL','DRUG','SMR',\
'DEPRESS','LBWT','EMERG','Attendance','Attainment','no_qualifications','not_participating',\
'University','drive_petrol','drive_GP','drive_post','drive_primary','drive_retail',\
'drive_secondary','PT_GP','PT_post','PT_retail','Broadband','crime_count','crime_rate',\
'overcrowded_count','nocentralheating_count','overcrowded_rate','nocentralheating_rate']
#What Indicators do you want to investigate?
for i in range(1, 36):
        enum_list_dict.append(i)
def print_indicators(dct):
    print("Indicators available:")
    for dict_of_factors, enum_list_dict in dct.items():
        print("{} ({})".format(enum_list_dict,dict_of_factors))
dict_of_factors= dict(zip(enum_list_dict, list_of_factors))
print_indicators(dict_of_factors)
#recieving user input
input_list = []
n = int(input("How many different Indicators do you wish to investigate for each postcode?\
 Type the number(s) corisponding to the Indicators you wish to for each postcode: "))

for i in range(0, n):
    ele = int(input())
    input_list.append(ele)
print("...")

#list generation
council_area_list=[]
total_population=[]
Working_Age_population=[]
Income_rate=[]
Income_count=[]
Employment_rate=[]
Employment_count=[]
CIF=[]
ALCOHOL=[]
DRUG=[]
SMR=[]
DEPRESS=[]
LBWT=[]
EMERG=[]
Attendance=[]
Attainment=[]
no_qualifications=[]
not_participating=[]
University=[]
drive_petrol=[]
drive_GP=[]
drive_post=[]
drive_primary=[]
drive_retail=[]
drive_secondary=[]
PT_GP=[]
PT_post=[]
PT_retail=[]
Broadband=[]
crime_count=[]
crime_rate=[]
overcrowded_count=[]
nocentralheating_count=[]
overcrowded_rate=[]
nocentralheating_rate=[]
requested_info_list=[]

factorsList = [council_area_list, total_population, Working_Age_population,\
 Income_rate, Income_count, Employment_rate, Employment_count, CIF, ALCOHOL, DRUG,\
 SMR, DEPRESS, LBWT, EMERG, Attendance, Attainment, no_qualifications, not_participating,\
 University, drive_petrol, drive_GP, drive_post, drive_primary, drive_retail, drive_secondary,\
 PT_GP,PT_post, PT_retail, Broadband, crime_count, crime_rate, overcrowded_count,\
 nocentralheating_count, overcrowded_rate, nocentralheating_rate, requested_info_list]
#excel load input
wb2=load_workbook('/Users/Thomas Clouston/Documents/py/SIMDQuerytool/SIMDQuery2016/00548707.xlsx')
SIMD_Indicators_Data= wb2['Data']
# add postcodes and datazone_lists to requested_info for output
requested_info_list.append(postcodes)
requested_info_list.append(datazone_lists)
for datazone in datazone_lists:
    for i in range(1,SIMD_Indicators_Data.max_row):
       if SIMD_Indicators_Data.cell(row=i, column=1).value == datazone:
          for j in range(37):
              if j in input_list:
                  factorName = factorsList[j]
                  if len(factorName) == 0:
                      factorName.append(list_of_factors[j])
                      factorName.append((str(SIMD_Indicators_Data.cell(row=i, column=j+2).value)))
                  else:
                      factorName.append((str(SIMD_Indicators_Data.cell(row=i, column=j+2).value)))
                  if factorName not in requested_info_list:
                      requested_info_list.append(factorName)

df = pd.DataFrame(requested_info_list[1:], columns=(requested_info_list[0]))
df1=df.transpose()
#Round each factor to set decimal place
df1.round({"1":1})
df1.to_excel(r"D:/Users/Thomas/Documents/py/SIMDQuerytool/SIMDQuery2016/Requested_info_for_inputed_postcodes.xlsx", sheet_name='Requested Data')
