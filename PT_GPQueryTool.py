from openpyxl import Workbook, load_workbook
import pandas as pd
#CSV input file:
df = pd.read_csv("/Users/Thomas/Documents/py/SIMDQuerytool/PT_GP/QueryPostcodes.csv")
#number of postcodes entered output
number_of_postcodes = int(len(df))
print ("Number of postcodes queried: ", number_of_postcodes)
print("Do NOT press any key...")
#change input csv to list
postcode_list= df.values.tolist()
#SIMD_2016_Postcode_lookup_tool return
path = "/Users/Thomas/Documents/py/SIMDQuerytool/PT_GP/2016Postcodes.xlsx"
wb = load_workbook(path)
ws= wb['All postcodes']
postcodes=[]
datazone_lists=[]
 # input logic
for i in range(1,ws.max_row):
    for postcode in postcode_list:
       if ws.cell(row=i, column=1).value == postcode[0]:
          for j in range(1,2):
              postcodes.append (str(ws.cell(row=i, column=j).value))
          for j in range(2,3):
              datazone_lists.append (str(ws.cell(row=i, column=j).value))

#excel load input
wb2=load_workbook('/Users/Thomas/Documents/py/SIMDQuerytool/PT_GP/00548707.xlsx')
SIMD_Indicators_Data= wb2['Data']
# add postcodes and requested_info for output
PT_GP=[]
for datazone in datazone_lists:
    for i in range(1,SIMD_Indicators_Data.max_row):
       if SIMD_Indicators_Data.cell(row=i, column=1).value == datazone:
                      PT_GP.append((str(SIMD_Indicators_Data.cell(row=i, column=28).value)))
#change from string to float
float_PT_GP=[]
for time in PT_GP:
    float_PT_GP.append(float(time))
#Round each factor to set decimal place
rounded_float_PT_GP= [round(time,1) for time in float_PT_GP]

#reorder to inputed format for output
dict_of_output_postcodes_and_PT_GP= dict(zip(postcodes,rounded_float_PT_GP))
list_of_postcodes=[]
for sublist in postcode_list:
    for item in sublist:
        list_of_postcodes.append(item)

outputdict={}
for postcode in list_of_postcodes:
    for postcodes,time in dict_of_output_postcodes_and_PT_GP.items():
            if postcode == postcodes:
                outputdict.update({postcodes:time})
#display and publish data
df = pd.DataFrame(list(outputdict.items()), columns= ['Postcodes','PT_GP'])
df.to_excel(r"D:/Users/Thomas/Documents/py/SIMDQuerytool/PT_GP/Requested_for_inputed_postcodes.xlsx", sheet_name='Requested Data')
